from openpyxl.utils import get_column_letter as get_col
from models import MeliItem
from datetime import datetime
from typing import List, Dict
import openpyxl as excel
import os, json

def itemsFromList(items_list: List[Dict]) -> List[MeliItem]:
    meli_items: List[MeliItem] = []
    for item in items_list:
        meli_items.append(MeliItem.fromDict(item))
    return meli_items
    
def itemsToExcel(items: List[MeliItem], extra_name="", save_path="./") -> str:
    assert os.path.isdir(save_path), "Save path does not exist"
    today_excel = excel.Workbook()
    today_excel.create_sheet("Items")
    
    for worksheet in today_excel.worksheets:
        if worksheet.title != "Items":
            today_excel.remove(worksheet)
    
    items_sheet = today_excel["Items"]
    
    headers:Dict[str,str] = MeliItem.getAttributes()
    
    items_sheet.append(headers)
    items_sheet.row_dimensions[0].alignment = excel.styles.Alignment(horizontal="center")
    items_sheet.row_dimensions[0].font = excel.styles.Font(bold=True)
    
    purple_fill = excel.styles.PatternFill(start_color="6600ff", end_color="6600ff", fill_type="solid")
    
    for h in range(1,len(headers)+1):
        items_sheet[f"{get_col(h)}1"].fill = purple_fill
        
    for item in items:
        print(item.Values)
        items_sheet.append(item.Values)
    
    file_name = f"{datetime.now().strftime('%Y-%m-%d-%H-%M')}_items.xlsx"
    file_name = extra_name + file_name
    today_file_name = os.path.join(save_path, file_name)
    today_excel.save(today_file_name)
    
    return file_name

def itemsToJson(items: List[MeliItem]) -> str:
    json_list = [item.toJson() for item in items]
    return json_list